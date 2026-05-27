"""
签到详情 xlsx → 标准零散时数表 xlsx
通用处理脚本，支持命令行参数。
"""
import argparse
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side


def extract_class(dept_str: str) -> str:
    if pd.isna(dept_str):
        return ""
    parts = str(dept_str).split("-")
    cls = parts[-1].strip() if parts else ""
    cls = re.sub(r'[\uff08(][^\uff09)]*[\uff09)]', '', cls)
    return cls


def extract_date_prefix(raw_df: pd.DataFrame) -> str:
    """从 Row 1 的时间字段提取 M.DD 格式的日期前缀"""
    time_str = str(raw_df.iloc[1, 1])
    match = re.search(r'(\d{4})-(\d{2})-(\d{2})', time_str)
    if match:
        month = str(int(match.group(2)))
        day = str(int(match.group(3)))
        return f"{month}.{day}"
    return ""


def process(input_path: str, output_path: str, activity_name: str,
            activity_date: str, hours: str):
    raw = pd.read_excel(input_path, header=None)

    data_rows = raw.iloc[10:].copy()
    data_rows.columns = range(data_rows.shape[1])

    signed = data_rows[
        data_rows[6].notna() & (data_rows[6].astype(str).str.strip() != "未签到")
    ].copy()

    result = pd.DataFrame({
        "姓名": signed[0].values,
        "班级": [extract_class(d) for d in signed[2].values],
        "学号": signed[4].values,
        "活动名称": activity_name,
        "活动日期": activity_date,
        "课时数": hours,
    })

    result["学号"] = result["学号"].astype(str).str.replace(r'\.0$', '', regex=True)
    result.sort_values("学号", inplace=True)
    result.reset_index(drop=True, inplace=True)

    # 写入 xlsx
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = ["姓名", "班级", "学号", "活动名称", "活动日期", "课时数"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, name="Arial", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for _, row in result.iterrows():
        ws.append([row["姓名"], row["班级"], str(row["学号"]),
                   row["活动名称"], row["活动日期"], row["课时数"]])

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in r:
            if cell.row > 1:
                cell.font = Font(name="Arial", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    col_widths = {"A": 14, "B": 14, "C": 18, "D": 20, "E": 14, "F": 16}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    wb.save(output_path)
    print(f"已生成: {output_path}")
    print(f"签到人数: {len(result)}")
    print(result.to_string(index=False))


def main():
    parser = argparse.ArgumentParser(description="签到详情 → 标准时数表")
    parser.add_argument("--input", required=True, help="签到详情 xlsx 路径")
    parser.add_argument("--output", required=True, help="输出 xlsx 路径")
    parser.add_argument("--activity-name", default=None, help="活动名称（--auto 可省略）")
    parser.add_argument("--activity-date", default=None, help="活动日期 YYYY.MM.DD（--auto 可省略）")
    parser.add_argument("--hours", default="就业指导2课时", help="课时数")
    parser.add_argument("--auto", action="store_true",
                        help="自动从源文件推断活动名称和日期")
    args = parser.parse_args()

    if args.auto:
        raw = pd.read_excel(args.input, header=None)
        date_prefix = extract_date_prefix(raw)
        topic = str(raw.iloc[0, 1]).strip()
        if not topic.startswith(date_prefix):
            args.activity_name = f"{date_prefix}{topic}"
        else:
            args.activity_name = topic
        time_str = str(raw.iloc[1, 1])
        match = re.search(r'(\d{4})-(\d{2})-(\d{2})', time_str)
        if match:
            args.activity_date = f"{match.group(1)}.{match.group(2)}.{match.group(3)}"

    if not args.activity_name or not args.activity_date:
        parser.error("需要 --activity-name 和 --activity-date，或使用 --auto 自动推断")

    process(args.input, args.output, args.activity_name,
            args.activity_date, args.hours)


if __name__ == "__main__":
    main()
