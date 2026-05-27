"""
签到详情 → 标准时数表：核心处理逻辑
独立于 GUI，可被 GUI / CLI 共同调用。
"""
import re
from pathlib import Path
from dataclasses import dataclass

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side


@dataclass
class CheckinResult:
    """处理结果"""
    success: bool
    count: int = 0
    output_path: str = ""
    message: str = ""
    preview: str = ""


def extract_class(dept_str) -> str:
    if pd.isna(dept_str):
        return ""
    parts = str(dept_str).split("-")
    cls = parts[-1].strip() if parts else ""
    cls = re.sub(r'[（(][^）)]*[）)]', '', cls)
    return cls


def extract_date_prefix(raw: pd.DataFrame) -> str:
    time_str = str(raw.iloc[1, 1])
    m = re.search(r'(\d{4})-(\d{2})-(\d{2})', time_str)
    if m:
        return f"{int(m.group(2))}.{int(m.group(3))}"
    return ""


def auto_detect(input_path: str) -> dict:
    """从源文件自动推断活动名称和日期，返回 dict"""
    raw = pd.read_excel(input_path, header=None)
    date_prefix = extract_date_prefix(raw)
    topic = str(raw.iloc[0, 1]).strip()

    if date_prefix and not topic.startswith(date_prefix):
        activity_name = f"{date_prefix}{topic}"
    else:
        activity_name = topic

    time_str = str(raw.iloc[1, 1])
    m = re.search(r'(\d{4})-(\d{2})-(\d{2})', time_str)
    activity_date = f"{m.group(1)}.{m.group(2)}.{m.group(3)}" if m else ""

    return {
        "activity_name": activity_name,
        "activity_date": activity_date,
        "topic": topic,
        "date_prefix": date_prefix,
    }


def process_checkin(input_path: str, output_path: str,
                    activity_name: str, activity_date: str,
                    hours: str, callback=None) -> CheckinResult:
    """
    主处理函数。
    callback: 可选的日志回调 callback(msg: str)
    """
    def log(msg):
        if callback:
            callback(msg)

    try:
        log(f"正在读取：{Path(input_path).name}")
        raw = pd.read_excel(input_path, header=None)

        data_rows = raw.iloc[10:].copy()
        data_rows.columns = range(data_rows.shape[1])

        total = len(data_rows)
        signed = data_rows[
            data_rows[6].notna() & (data_rows[6].astype(str).str.strip() != "未签到")
        ].copy()

        unsigned_count = total - len(signed)
        log(f"总人数：{total}，已签到：{len(signed)}，未签到：{unsigned_count}")

        result = pd.DataFrame({
            "姓名": signed[0].values,
            "班级": [extract_class(d) for d in signed[2].values],
            "学号": signed[4].values,
            "活动名称": activity_name,
            "活动日期": activity_date,
            "课时数": hours,
        })

        result["学号"] = result["学号"].astype(str).str.replace(r'\.0$', '', regex=True)
        result.sort_values("学号", inplace=True)
        result.reset_index(drop=True, inplace=True)

        # 写入 xlsx
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        headers = ["姓名", "班级", "学号", "活动名称", "活动日期", "课时数"]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True, name="Arial", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for _, row in result.iterrows():
            ws.append([row["姓名"], row["班级"], str(row["学号"]),
                       row["活动名称"], row["活动日期"], row["课时数"]])

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
        )

        for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
            for cell in r:
                if cell.row > 1:
                    cell.font = Font(name="Arial", size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

        col_widths = {"A": 14, "B": 14, "C": 18, "D": 20, "E": 14, "F": 16}
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w

        wb.save(output_path)
        log(f"✅ 已保存：{Path(output_path).name}")

        preview = result.head(5).to_string(index=False)
        return CheckinResult(
            success=True,
            count=len(result),
            output_path=output_path,
            message=f"成功处理 {len(result)} 条签到记录",
            preview=preview,
        )

    except Exception as e:
        log(f"❌ 处理失败：{e}")
        return CheckinResult(success=False, message=str(e))
