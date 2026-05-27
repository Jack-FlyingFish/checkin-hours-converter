"""
签到详情 → 时数表转换工具 v2.0
单文件完整版，双击即可运行。
"""
import re
import sys
import threading
import warnings
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass

# 忽略 openpyxl 加载非标准 Excel 文件时产生的样式丢失警告
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

import customtkinter as ctk
from tkinter import filedialog, END


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  核心处理逻辑
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


@dataclass
class CheckinResult:
    success: bool
    count: int = 0
    output_path: str = ""
    message: str = ""


def _extract_class(dept_str) -> str:
    if pd.isna(dept_str):
        return ""
    parts = str(dept_str).split("-")
    cls = parts[-1].strip() if parts else ""
    cls = re.sub(r'[（(][^）)]*[）)]', '', cls)
    return cls


def _extract_date_prefix(raw: pd.DataFrame) -> str:
    time_str = str(raw.iloc[1, 1])
    m = re.search(r'(\d{4})-(\d{2})-(\d{2})', time_str)
    if m:
        return f"{int(m.group(2))}.{int(m.group(3))}"
    return ""


def sanitize_filename(filename: str) -> str:
    """清除 Windows 文件路径中的非法字符"""
    return re.sub(r'[\\/:*?"<>|]', '', filename)


def auto_detect(input_path: str) -> dict:
    raw = pd.read_excel(input_path, header=None)
    date_prefix = _extract_date_prefix(raw)
    topic = str(raw.iloc[0, 1]).strip()

    if date_prefix and not topic.startswith(date_prefix):
        activity_name = f"{date_prefix}{topic}"
    else:
        activity_name = topic
    # 确保日期前缀与活动名称之间无空格，如 "3.11海亮校园宣讲"
    if date_prefix:
        activity_name = activity_name.replace(date_prefix + " ", date_prefix)
    
    activity_name = sanitize_filename(activity_name)

    time_str = str(raw.iloc[1, 1])
    m = re.search(r'(\d{4})-(\d{2})-(\d{2})', time_str)
    activity_date = f"{m.group(1)}.{m.group(2)}.{m.group(3)}" if m else ""

    return {
        "activity_name": activity_name,
        "activity_date": activity_date,
    }


def process_checkin(input_path: str, output_path: str,
                    activity_name: str, activity_date: str,
                    hours: str, callback=None) -> CheckinResult:
    def log(msg):
        if callback:
            callback(msg)

    try:
        log(f"正在读取：{Path(input_path).name}")
        raw = pd.read_excel(input_path, header=None)

        data_rows = raw.iloc[10:].copy()
        data_rows.columns = range(data_rows.shape[1])

        total = len(data_rows)

        # 筛选：已签到 且 学号非隐藏
        signed = data_rows[
            data_rows[6].notna()
            & (data_rows[6].astype(str).str.strip() != "未签到")
            & (data_rows[4].astype(str).str.strip() != "已隐藏")
        ].copy()

        unsigned = total - len(signed)
        log(f"总人数：{total}，有效签到：{len(signed)}，排除：{unsigned}")

        result = pd.DataFrame({
            "姓名": signed[0].values,
            "班级": [_extract_class(d) for d in signed[2].values],
            "学号": signed[4].values,
            "活动名称": activity_name,
            "活动日期": activity_date,
            "课时数": hours,
        })

        result["学号"] = result["学号"].astype(str).str.replace(r'\.0$', '', regex=True)
        # 先按班级排序，同班级内再按学号升序
        result.sort_values(["班级", "学号"], inplace=True)
        result.reset_index(drop=True, inplace=True)

        # ── 写入 xlsx ──
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        headers = ["姓名", "班级", "学号", "活动名称", "活动日期", "课时数"]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True, name="Arial", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for _, row in result.iterrows():
            ws.append([row["姓名"], row["班级"], str(row["学号"]),
                       row["活动名称"], row["活动日期"], row["课时数"]])

        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
        )
        for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
            for cell in r:
                if cell.row > 1:
                    cell.font = Font(name="Arial", size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        for col, w in {"A": 14, "B": 14, "C": 18, "D": 20, "E": 14, "F": 16}.items():
            ws.column_dimensions[col].width = w

        wb.save(output_path)
        log(f"✅ 已保存：{Path(output_path).name}")

        return CheckinResult(True, len(result), output_path,
                             f"成功处理 {len(result)} 条签到记录")

    except Exception as e:
        log(f"❌ 处理失败：{e}")
        return CheckinResult(False, message=str(e))


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  GUI
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

ctk.set_default_color_theme("blue")

FONT_TITLE = ("Microsoft YaHei UI", 18, "bold")
FONT_LABEL = ("Microsoft YaHei UI", 13)
FONT_ENTRY = ("Microsoft YaHei UI", 12)
FONT_BTN   = ("Microsoft YaHei UI", 13, "bold")
FONT_LOG   = ("Consolas", 11)
FONT_SMALL = ("Microsoft YaHei UI", 11)

ACCENT       = "#3B8ED0"
SUCCESS      = "#2FA572"
ERROR        = "#E04040"

HOUR_TYPES   = ["就业指导", "创新创业", "自定义"]
HOUR_NUMS    = ["1", "2", "3", "4", "自定义"]


class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("签到详情 → 时数表转换工具")
        self.geometry("660x720")
        self.minsize(600, 680)

        self._input_path = ""
        self._auto_info = {}

        self._build_ui()

    # ─── 构建 ─────────────────────────────────────────────
    def _build_ui(self):
        self.grid_columnconfigure(0, weight=1)

        # 顶部栏：标题 + 主题切换
        top = ctk.CTkFrame(self, fg_color="transparent")
        top.grid(row=0, column=0, padx=20, pady=(16, 4), sticky="ew")
        top.grid_columnconfigure(0, weight=1)

        title_box = ctk.CTkFrame(top, fg_color="transparent")
        title_box.grid(row=0, column=0, sticky="w")
        ctk.CTkLabel(title_box, text="📋 签到详情 → 时数表", font=FONT_TITLE
                     ).pack(anchor="w")
        ctk.CTkLabel(title_box, text="将签到平台导出的 xlsx 转换为标准零散时数统计表",
                     font=FONT_SMALL, text_color="gray60"
                     ).pack(anchor="w", pady=(2, 0))

        # 主题切换
        theme_box = ctk.CTkFrame(top, fg_color="transparent")
        theme_box.grid(row=0, column=1, sticky="ne", padx=(8, 0))
        self.theme_var = ctk.StringVar(value="dark")
        self.theme_seg = ctk.CTkSegmentedButton(
            theme_box, values=["☀️ 亮", "🌙 暗"],
            font=FONT_SMALL, width=100,
            command=self._on_theme,
        )
        self.theme_seg.set("🌙 暗")
        self.theme_seg.pack(pady=(4, 0))

        # ── 文件区 ──
        ff = ctk.CTkFrame(self)
        ff.grid(row=1, column=0, padx=20, pady=(10, 6), sticky="ew")
        ff.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(ff, text="📂 签到文件", font=FONT_LABEL, width=90
                     ).grid(row=0, column=0, padx=(14, 6), pady=(14, 6), sticky="w")
        self.input_entry = ctk.CTkEntry(ff, font=FONT_ENTRY,
                                        placeholder_text="选择签到详情 .xlsx 文件…")
        self.input_entry.grid(row=0, column=1, padx=4, pady=(14, 6), sticky="ew")
        ctk.CTkButton(ff, text="浏览", width=64, font=FONT_SMALL,
                      command=self._browse_input
                      ).grid(row=0, column=2, padx=(4, 14), pady=(14, 6))

        ctk.CTkLabel(ff, text="📁 输出路径", font=FONT_LABEL, width=90
                     ).grid(row=1, column=0, padx=(14, 6), pady=(6, 14), sticky="w")
        self.output_entry = ctk.CTkEntry(ff, font=FONT_ENTRY,
                                         placeholder_text="自动生成，或手动选择…")
        self.output_entry.grid(row=1, column=1, padx=4, pady=(6, 14), sticky="ew")
        ctk.CTkButton(ff, text="浏览", width=64, font=FONT_SMALL,
                      command=self._browse_output
                      ).grid(row=1, column=2, padx=(4, 14), pady=(6, 14))

        # ── 模式选择 ──
        mode_frame = ctk.CTkFrame(self)
        mode_frame.grid(row=2, column=0, padx=20, pady=6, sticky="ew")
        mode_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(mode_frame, text="模式", font=FONT_LABEL, width=50
                     ).grid(row=0, column=0, padx=(14, 6), pady=12, sticky="w")
        self.mode_var = ctk.StringVar(value="auto")
        self.mode_seg = ctk.CTkSegmentedButton(
            mode_frame, values=["🔄 自动推断", "✏️ 手动填写"],
            font=FONT_SMALL, command=self._on_mode,
        )
        self.mode_seg.set("🔄 自动推断")
        self.mode_seg.grid(row=0, column=1, padx=4, pady=12, sticky="w")

        # ── 时数设置 ──
        hf = ctk.CTkFrame(self)
        hf.grid(row=3, column=0, padx=20, pady=6, sticky="ew")
        hf.grid_columnconfigure(3, weight=1)

        ctk.CTkLabel(hf, text="时数", font=FONT_LABEL, width=50
                     ).grid(row=0, column=0, padx=(14, 6), pady=12, sticky="w")

        # ── 类型插槽（下拉 或 自定义输入+取消）──
        self.type_slot = ctk.CTkFrame(hf, fg_color="transparent")
        self.type_slot.grid(row=0, column=1, padx=4, pady=12)

        self.hour_type_var = ctk.StringVar(value="就业指导")
        self.hour_type_menu = ctk.CTkOptionMenu(
            self.type_slot, values=HOUR_TYPES, variable=self.hour_type_var,
            width=110, font=FONT_SMALL, command=self._on_hour_type,
        )
        self.hour_type_menu.pack(side="left")

        self.custom_type_entry = ctk.CTkEntry(
            self.type_slot, font=FONT_ENTRY, width=88, placeholder_text="类型名称"
        )
        self.type_cancel_btn = ctk.CTkButton(
            self.type_slot, text="×", width=26, height=28, font=("Arial", 13),
            fg_color="transparent", hover_color=("gray80", "gray30"),
            text_color=("gray50", "gray60"), border_width=0,
            command=self._cancel_custom_type,
        )

        # ── 数量插槽──
        self.num_slot = ctk.CTkFrame(hf, fg_color="transparent")
        self.num_slot.grid(row=0, column=2, padx=4, pady=12)

        self.hour_num_var = ctk.StringVar(value="2")
        self.hour_num_menu = ctk.CTkOptionMenu(
            self.num_slot, values=HOUR_NUMS, variable=self.hour_num_var,
            width=72, font=FONT_SMALL, command=self._on_hour_num,
        )
        self.hour_num_menu.pack(side="left")

        self.custom_num_entry = ctk.CTkEntry(
            self.num_slot, font=FONT_ENTRY, width=50, placeholder_text="数字"
        )
        self.num_cancel_btn = ctk.CTkButton(
            self.num_slot, text="×", width=26, height=28, font=("Arial", 13),
            fg_color="transparent", hover_color=("gray80", "gray30"),
            text_color=("gray50", "gray60"), border_width=0,
            command=self._cancel_custom_num,
        )

        ctk.CTkLabel(hf, text="课时", font=FONT_LABEL
                     ).grid(row=0, column=3, padx=(2, 4), pady=12, sticky="w")

        # 预览标签
        self.hours_preview = ctk.CTkLabel(
            hf, text="→ 就业指导2课时", font=FONT_SMALL, text_color="gray60"
        )
        self.hours_preview.grid(row=0, column=4, padx=(8, 14), pady=12, sticky="e")

        # ── 手动参数区 ──
        self.manual_frame = ctk.CTkFrame(self)
        self.manual_frame.grid(row=4, column=0, padx=20, pady=6, sticky="ew")
        self.manual_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(self.manual_frame, text="活动名称", font=FONT_LABEL, width=80
                     ).grid(row=0, column=0, padx=(14, 6), pady=(14, 6), sticky="w")
        self.name_entry = ctk.CTkEntry(self.manual_frame, font=FONT_ENTRY,
                                       placeholder_text="如：3.11海亮校园宣讲")
        self.name_entry.grid(row=0, column=1, padx=(4, 14), pady=(14, 6), sticky="ew")

        ctk.CTkLabel(self.manual_frame, text="活动日期", font=FONT_LABEL, width=80
                     ).grid(row=1, column=0, padx=(14, 6), pady=(6, 14), sticky="w")
        self.date_entry = ctk.CTkEntry(self.manual_frame, font=FONT_ENTRY,
                                       placeholder_text="如：2026.03.11")
        self.date_entry.grid(row=1, column=1, padx=(4, 14), pady=(6, 14), sticky="ew")

        self._set_manual_state(False)

        # ── 转换按钮 ──
        self.convert_btn = ctk.CTkButton(
            self, text="▶  开始转换", font=FONT_BTN, height=44,
            corner_radius=10,
        )
        self.convert_btn.grid(row=5, column=0, padx=20, pady=(12, 6), sticky="ew")
        # 按下抬起后才触发
        self.convert_btn.bind("<ButtonRelease-1>", self._on_convert_release)
        self.convert_btn.configure(command=lambda: None)  # 禁用 ctk 默认点击

        # ── 日志面板 ──
        ctk.CTkLabel(self, text="处理日志", font=FONT_SMALL,
                     text_color="gray60", anchor="w"
                     ).grid(row=6, column=0, padx=24, pady=(10, 2), sticky="w")

        self.log_box = ctk.CTkTextbox(self, font=FONT_LOG, height=160,
                                      state="disabled", wrap="word",
                                      corner_radius=8)
        self.log_box.grid(row=7, column=0, padx=20, pady=(0, 6), sticky="nsew")
        self.grid_rowconfigure(7, weight=1)

        # ── 状态栏 ──
        self.status_label = ctk.CTkLabel(
            self, text="就绪  |  就业指导2课时", font=FONT_SMALL,
            text_color="gray50", anchor="w",
        )
        self.status_label.grid(row=8, column=0, padx=24, pady=(2, 12), sticky="w")

        ctk.CTkLabel(
            self, text="Xulehan Yu", font=("Microsoft YaHei UI", 10),
            text_color="gray35", anchor="e",
        ).grid(row=8, column=0, padx=24, pady=(2, 12), sticky="e")

        # 初始主题
        ctk.set_appearance_mode("dark")

    # ─── 主题 ─────────────────────────────────────────────
    def _on_theme(self, value):
        mode = "light" if "亮" in value else "dark"
        ctk.set_appearance_mode(mode)

    # ─── 模式 ─────────────────────────────────────────────
    def _on_mode(self, value):
        is_manual = "手动" in value
        self.mode_var.set("manual" if is_manual else "auto")
        self._set_manual_state(is_manual)
        if not is_manual and self._input_path:
            self._auto_parse()

    def _set_manual_state(self, enabled: bool):
        state = "normal" if enabled else "disabled"
        self.name_entry.configure(state=state)
        self.date_entry.configure(state=state)
        self.manual_frame.configure(
            fg_color=("gray92", "gray14") if enabled else ("gray88", "gray20")
        )

    # ─── 时数 ─────────────────────────────────────────────
    def _on_hour_type(self, value):
        if value == "自定义":
            self.hour_type_menu.pack_forget()
            self.custom_type_entry.pack(side="left")
            self.type_cancel_btn.pack(side="left", padx=(2, 0))
            self.custom_type_entry.focus()
            self.custom_type_entry.bind("<Return>", lambda e: self._refresh_hours())
            self.custom_type_entry.bind("<FocusOut>", lambda e: self._refresh_hours())
        self._refresh_hours()

    def _cancel_custom_type(self):
        self.custom_type_entry.pack_forget()
        self.type_cancel_btn.pack_forget()
        self.hour_type_var.set("就业指导")
        self.hour_type_menu.pack(side="left")
        self._refresh_hours()

    def _on_hour_num(self, value):
        if value == "自定义":
            self.hour_num_menu.pack_forget()
            self.custom_num_entry.pack(side="left")
            self.num_cancel_btn.pack(side="left", padx=(2, 0))
            self.custom_num_entry.focus()
            self.custom_num_entry.bind("<Return>", lambda e: self._refresh_hours())
            self.custom_num_entry.bind("<FocusOut>", lambda e: self._refresh_hours())
        self._refresh_hours()

    def _cancel_custom_num(self):
        self.custom_num_entry.pack_forget()
        self.num_cancel_btn.pack_forget()
        self.hour_num_var.set("2")
        self.hour_num_menu.pack(side="left")
        self._refresh_hours()

    def _get_hour_type(self) -> str:
        v = self.hour_type_var.get()
        if v == "自定义":
            return self.custom_type_entry.get().strip() or "自定义"
        return v

    def _get_hour_num(self) -> str:
        v = self.hour_num_var.get()
        if v == "自定义":
            return self.custom_num_entry.get().strip() or "?"
        return v

    def _get_hours_str(self) -> str:
        return f"{self._get_hour_type()}{self._get_hour_num()}课时"

    def _refresh_hours(self):
        h = self._get_hours_str()
        self.hours_preview.configure(text=f"→ {h}")
        self.status_label.configure(text=f"就绪  |  {h}", text_color="gray50")

    # ─── 文件 ─────────────────────────────────────────────
    def _browse_input(self):
        path = filedialog.askopenfilename(
            title="选择签到详情文件",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
        )
        if not path:
            return
        self._input_path = path
        self.input_entry.delete(0, END)
        self.input_entry.insert(0, path)
        if self.mode_var.get() == "auto":
            self._auto_parse()

    def _browse_output(self):
        path = filedialog.asksaveasfilename(
            title="保存输出文件",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
        )
        if path:
            self.output_entry.delete(0, END)
            self.output_entry.insert(0, path)

    def _auto_parse(self):
        if not self._input_path:
            return
        try:
            info = auto_detect(self._input_path)
            self._auto_info = info

            self.name_entry.configure(state="normal")
            self.name_entry.delete(0, END)
            self.name_entry.insert(0, info["activity_name"])
            self.date_entry.configure(state="normal")
            self.date_entry.delete(0, END)
            self.date_entry.insert(0, info["activity_date"])

            if self.mode_var.get() == "auto":
                self.name_entry.configure(state="disabled")
                self.date_entry.configure(state="disabled")

            out = str(Path(self._input_path).parent / f"{info['activity_name']}.xlsx")
            self.output_entry.delete(0, END)
            self.output_entry.insert(0, out)

            self._log(f"已解析：{info['activity_name']}  |  {info['activity_date']}")
        except Exception as e:
            self._log(f"⚠ 解析失败：{e}")

    # ─── 日志 ─────────────────────────────────────────────
    def _log(self, msg: str):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_box.configure(state="normal")
        self.log_box.insert(END, f"[{ts}] {msg}\n")
        self.log_box.see(END)
        self.log_box.configure(state="disabled")

    # ─── 转换 ─────────────────────────────────────────────
    def _on_convert_release(self, event=None):
        # 确保鼠标在按钮范围内释放
        w = self.convert_btn
        x, y = event.x, event.y
        if x < 0 or y < 0 or x > w.winfo_width() or y > w.winfo_height():
            return
        if str(w.cget("state")) == "disabled":
            return
        self._do_convert()

    def _do_convert(self):
        inp = self.input_entry.get().strip()
        if not inp or not Path(inp).is_file():
            self._log("❌ 请先选择有效的签到详情文件")
            return

        out = self.output_entry.get().strip()
        if not out:
            self._log("❌ 请指定输出路径")
            return

        if self.mode_var.get() == "manual":
            name = self.name_entry.get().strip()
            date = self.date_entry.get().strip()
            if not name or not date:
                self._log("❌ 手动模式下请填写活动名称和日期")
                return
        else:
            name = self._auto_info.get("activity_name", "")
            date = self._auto_info.get("activity_date", "")
            if not name:
                self._log("❌ 自动推断失败，请切换到手动模式")
                return

        hours = self._get_hours_str()

        self.convert_btn.configure(state="disabled", text="⏳ 处理中…")
        self.status_label.configure(text="正在处理…", text_color=ACCENT)

        def worker():
            r = process_checkin(inp, out, name, date, hours,
                                callback=lambda m: self.after(0, self._log, m))
            self.after(0, self._done, r)

        threading.Thread(target=worker, daemon=True).start()

    def _done(self, r: CheckinResult):
        self.convert_btn.configure(state="normal", text="▶  开始转换")
        if r.success:
            self.status_label.configure(
                text=f"✅ 完成  |  {r.count} 条  |  {Path(r.output_path).name}",
                text_color=SUCCESS,
            )
            self._log(f"共 {r.count} 条记录")
        else:
            self.status_label.configure(
                text=f"❌ 失败：{r.message}", text_color=ERROR,
            )


if __name__ == "__main__":
    App().mainloop()
